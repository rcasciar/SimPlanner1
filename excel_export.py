"""
Modulo per l'esportazione delle programmazioni in formato Excel.
"""
import io
import pandas as pd
from datetime import datetime
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models import ScheduleData


def create_excel_from_schedule(schedule_data, title="Programmazione Laboratori"):
    """
    Crea un file Excel dalla programmazione dei laboratori.
    
    Args:
        schedule_data: DataFrame o ScheduleData contenente la programmazione
        title: Titolo del documento
    
    Returns:
        BytesIO contenente il file Excel
    """
    # Verifica se schedule_data è un DataFrame o un oggetto ScheduleData
    if isinstance(schedule_data, pd.DataFrame):
        df = schedule_data
        if df.empty:
            return None
    else:
        # Se è un oggetto ScheduleData
        if hasattr(schedule_data, 'scheduled_labs') and not schedule_data.scheduled_labs:
            return None
        elif not hasattr(schedule_data, 'scheduled_labs'):
            # Se non ha l'attributo scheduled_labs, potrebbe essere un altro tipo di oggetto
            try:
                # Prova a convertirlo in DataFrame direttamente
                df = pd.DataFrame(schedule_data)
                if df.empty:
                    return None
            except:
                return None
        else:
            # Get data frame with schedule info
            df = schedule_data.get_data_frame()
            
            if df.empty:
                return None
    
    # Create Excel workbook
    wb = Workbook()
    
    # Use the default worksheet for the main schedule
    ws = wb.active
    ws.title = "Programmazione"
    
    # Set the title
    ws['A1'] = "Programmazione Laboratori"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Create header row - starting from row 3
    header_row = 3
    headers = ["Giorno", "Fascia Oraria", "Laboratorio", "Aule", "Gruppi"]
    
    for idx, header in enumerate(headers):
        ws.cell(row=header_row, column=idx+1).value = header
        ws.cell(row=header_row, column=idx+1).font = Font(bold=True)
    
    # Style the header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Giorno
    ws.column_dimensions['B'].width = 15  # Fascia Oraria
    ws.column_dimensions['C'].width = 30  # Laboratorio
    ws.column_dimensions['D'].width = 30  # Aule
    ws.column_dimensions['E'].width = 25  # Gruppi
    
    # Prepare data
    # Sort by day, then start time
    # Controllo se le colonne esistono nel DataFrame e adatta gli ordinamenti
    if 'Day' in df.columns and 'Start' in df.columns:
        df = df.sort_values(by=['Day', 'Start'])
    elif 'data' in df.columns and 'ora_inizio' in df.columns:
        df = df.sort_values(by=['data', 'ora_inizio'])
    
    current_row = header_row + 1
    lunch_break_added = set()  # Keep track of days where lunch break has been added
    
    # Group by day - adatta in base alle colonne disponibili
    if 'Day' in df.columns:
        days = sorted(df['Day'].unique())
    elif 'data' in df.columns:
        days = sorted(df['data'].unique())
    
    for day in days:
        # Log per debug
        print(f"Elaborando il giorno: {day} di tipo {type(day)}")
        
        # Adatta alla struttura delle colonne
        if 'Day' in df.columns:
            day_df = df[df['Day'] == day]
        elif 'data' in df.columns:
            day_df = df[df['data'] == day]
        
        # Fixed time slots
        time_slots = [
            {'start': '08:30', 'end': '11:00'},
            {'start': '11:10', 'end': '13:40'},
            {'start': '14:10', 'end': '17:10'}
        ]
        
        for time_slot in time_slots:
            slot_start = time_slot['start']
            slot_end = time_slot['end']
            # Adatta in base alle colonne disponibili
            if 'Start' in day_df.columns:
                time_df = day_df[day_df['Start'] == slot_start]
            elif 'ora_inizio' in day_df.columns:
                time_df = day_df[day_df['ora_inizio'] == slot_start]
            else:
                time_df = pd.DataFrame()  # DataFrame vuoto se non ci sono colonne compatibili
            
            if time_df.empty:
                continue
                
            # Group labs by name to show all rooms used for the same lab
            lab_groups = {}
            for _, row in time_df.iterrows():
                # Adatta in base alle colonne disponibili
                if 'Lab' in row:
                    lab_name = row['Lab']
                elif 'laboratorio' in row:
                    lab_name = row['laboratorio']
                else:
                    continue  # Salta la riga se non ha informazioni sul laboratorio
                
                # Adatta in base alle colonne disponibili
                if 'Room' in row:
                    room_name = row['Room']
                elif 'aula' in row:
                    room_name = row['aula']
                else:
                    room_name = "Aula non specificata"
                
                if lab_name not in lab_groups:
                    lab_groups[lab_name] = []
                lab_groups[lab_name].append(room_name)
            
            for lab_name, rooms in lab_groups.items():
                # Find all student groups for this lab - solo per oggetti ScheduleData
                lab_student_groups = []
                
                # Se schedule_data è un oggetto ScheduleData, estrai i gruppi
                if not isinstance(schedule_data, pd.DataFrame) and hasattr(schedule_data, 'scheduled_labs'):
                    for sched_lab in schedule_data.scheduled_labs:
                        if (sched_lab.lab.name == lab_name and 
                            sched_lab.time_slot.day == day and 
                            sched_lab.time_slot.start_time.strftime("%H:%M") == slot_start):
                            # Convert student IDs to group names
                            for student_id in sched_lab.students:
                                group_name = f"Gruppo {student_id+1}"
                                if group_name not in lab_student_groups:
                                    lab_student_groups.append(group_name)
                
                # Se è un DataFrame, estrai i gruppi dalla colonna 'Gruppo'
                elif isinstance(schedule_data, pd.DataFrame):
                    # Adatta le colonne in base ai nomi disponibili
                    if 'gruppo' in df.columns:
                        # Adatta anche le condizioni di filtro in base alle colonne del DataFrame
                        filter_conditions = []
                        if 'Day' in df.columns and 'Lab' in df.columns and 'Start' in df.columns:
                            lab_df = df[(df['Day'] == day) & (df['Lab'] == lab_name) & (df['Start'] == slot_start)]
                        elif 'data' in df.columns and 'laboratorio' in df.columns and 'ora_inizio' in df.columns:
                            lab_df = df[(df['data'] == day) & (df['laboratorio'] == lab_name) & (df['ora_inizio'] == slot_start)]
                        else:
                            lab_df = pd.DataFrame()  # DataFrame vuoto se non ci sono colonne compatibili
                        
                        if not lab_df.empty:
                            for gruppo in lab_df['gruppo'].unique():
                                if gruppo and str(gruppo) not in lab_student_groups:
                                    lab_student_groups.append(str(gruppo))
                
                # Add row for this lab - handling day as string or number
                try:
                    # Se day è un numero, usa day+1
                    if isinstance(day, int):
                        ws.cell(row=current_row, column=1).value = f"Giorno {day+1}"
                    else:
                        # Altrimenti usa il valore direttamente (già in formato "Giorno X" o data)
                        ws.cell(row=current_row, column=1).value = str(day)
                except Exception as e:
                    print(f"Errore nel formattare il giorno {day}: {str(e)}")
                    ws.cell(row=current_row, column=1).value = str(day)
                    
                ws.cell(row=current_row, column=2).value = f"{slot_start}–{slot_end}"
                
                # Get the duration from any row with this lab
                try:
                    # Determina le colonne in base al DataFrame
                    if 'Lab' in time_df.columns and 'Duration' in time_df.columns:
                        duration_row = time_df[time_df['Lab'] == lab_name].iloc[0]
                        duration = duration_row['Duration']
                    elif 'laboratorio' in time_df.columns:
                        duration_row = time_df[time_df['laboratorio'] == lab_name].iloc[0]
                        # Calcola la durata dalla differenza tra ora_inizio e ora_fine
                        if 'ora_inizio' in duration_row and 'ora_fine' in duration_row:
                            # Converti le stringhe in datetime per il calcolo
                            from datetime import datetime
                            ora_inizio = datetime.strptime(duration_row['ora_inizio'], '%H:%M')
                            ora_fine = datetime.strptime(duration_row['ora_fine'], '%H:%M')
                            # Calcola la differenza in minuti
                            diff = ora_fine - ora_inizio
                            duration = diff.seconds / 60
                        else:
                            duration = 150  # Default duration in minutes
                    else:
                        duration = 150  # Default duration in minutes
                    
                    ws.cell(row=current_row, column=3).value = f"{lab_name} ({duration/60:.1f} ore)"
                except (IndexError, KeyError):
                    # Se non riesce a trovare la durata, la imposta a un valore predefinito
                    ws.cell(row=current_row, column=3).value = f"{lab_name}"
                
                # Rooms
                ws.cell(row=current_row, column=4).value = ", ".join(rooms)
                
                # Student groups
                ws.cell(row=current_row, column=5).value = ", ".join(lab_student_groups)
                
                # Add borders to all cells
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                current_row += 1
            
            # Add lunch break after the second time slot if we haven't already for this day
            if time_slot['start'] == '11:10' and day not in lunch_break_added:
                lunch_break_added.add(day)
                
                # Consistent handling of day format per pausa pranzo
                try:
                    if isinstance(day, int):
                        ws.cell(row=current_row, column=1).value = f"Giorno {day+1}"
                    else:
                        # Altrimenti usa il valore direttamente
                        ws.cell(row=current_row, column=1).value = str(day)
                except Exception as e:
                    print(f"Errore nel formattare il giorno {day} per pausa pranzo: {str(e)}")
                    ws.cell(row=current_row, column=1).value = str(day)
                    
                ws.cell(row=current_row, column=2).value = "13:40–14:10"
                ws.cell(row=current_row, column=3).value = "PAUSA PRANZO"
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
                
                # Style the lunch break
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    if col == 3:  # The merged cell
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    else:
                        cell.alignment = Alignment(vertical='center')
                
                current_row += 1
        
        # Add a blank row after each day
        current_row += 1
    
    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_schedule_excel(schedule_data, filename="programmazione_laboratori.xlsx"):
    """
    Esporta la programmazione dei laboratori in un file Excel.
    
    Args:
        schedule_data: ScheduleData contenente la programmazione
        filename: Nome del file Excel
    
    Returns:
        BytesIO contenente il file Excel
    """
    try:
        # Log dettagliato per debug
        print(f"Tipo di schedule_data: {type(schedule_data)}")
        
        if isinstance(schedule_data, pd.DataFrame):
            print(f"Colonne nel DataFrame: {schedule_data.columns.tolist()}")
            print(f"Numero di righe: {len(schedule_data)}")
        
        # Prova a creare il file Excel
        excel_data = create_excel_from_schedule(schedule_data)
        
        if excel_data is None:
            print("create_excel_from_schedule ha restituito None")
            return None
        
        return excel_data
    except Exception as e:
        # Cattura e logga l'errore
        error_trace = traceback.format_exc()
        print(f"Errore in export_schedule_excel: {str(e)}")
        print(f"Traceback: {error_trace}")
        # Riprova con un approccio più semplice
        try:
            # Se schedule_data è un DataFrame, prova un'esportazione diretta
            if isinstance(schedule_data, pd.DataFrame):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    schedule_data.to_excel(writer, sheet_name='Programmazione', index=False)
                output.seek(0)
                return output
            return None
        except Exception as e2:
            print(f"Anche il fallback è fallito: {str(e2)}")
            return None